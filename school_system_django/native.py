import csv
import math
import re
from datetime import date
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone


IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")

DEFAULT_PRIMARY_KEYS = {
    "academic_year": "academic_id",
    "attendance_records": "attendance_id",
    "audit_log": "audit_id",
    "balance_adjustments": "adjustment_id",
    "class_timetable_entries": "timetable_id",
    "classes": "class_id",
    "communication_log": "communication_id",
    "database_backups_log": "backup_id",
    "e_learning_assignments": "assignment_id",
    "e_learning_notes": "note_id",
    "e_learning_submissions": "submission_id",
    "exam_sessions": "exam_id",
    "expenses": "expense_id",
    "fees_structure": "fee_id",
    "grades": "grade_id",
    "guardians": "guardian_id",
    "inventory_items": "item_id",
    "inventory_movements": "movement_id",
    "library_books": "book_id",
    "library_issues": "issue_id",
    "library_issues": "issue_id",
    "master_receipts": "master_receipt_id",
    "offline_sync_events": "event_id",
    "online_payment_requests": "request_id",
    "payment_allocations": "allocation_id",
    "payments": "payment_id",
    "portal_update_events": "event_id",
    "pos_sale_items": "sale_item_id",
    "pos_sales": "sale_id",
    "pupil_fee_overrides": "override_id",
    "pupils": "pupil_id",
    "receipts": "receipt_id",
    "result_entries": "entry_id",
    "result_sheets": "result_id",
    "school_settings": "setting_id",
    "student_performance_predictions": "prediction_id",
    "subjects": "subject_id",
    "teacher_attendance_records": "attendance_id",
    "teacher_profiles": "profile_id",
    "term_bills": "bill_id",
    "textbook_loans": "loan_id",
    "users": "user_id",
    "website_announcements": "announcement_id",
}

READ_ONLY_COLUMNS = {
    "created_at",
    "updated_at",
    "marked_at",
    "generated_at",
    "uploaded_at",
    "submitted_at",
    "synced_at",
    "published_at",
    "approved_at",
    "reviewed_at",
    "paid_at",
}

SYSTEM_USER_COLUMNS = {
    "created_by",
    "recorded_by",
    "marked_by",
    "generated_by",
    "uploaded_by",
    "updated_by",
    "published_by",
    "issued_by",
    "returned_by",
    "cleared_by",
}

DATE_DEFAULT_COLUMNS = {
    "admission_date",
    "attendance_date",
    "payment_date",
    "expense_date",
    "borrowed_date",
    "issue_date",
    "return_date",
    "due_date",
    "entry_date",
    "batch_date",
    "billed_on",
    "issued_date",
    "sale_date",
    "movement_date",
    "status_changed_on",
    "completed_on",
    "cleared_date",
}


def assert_identifier(name):
    if not IDENTIFIER_RE.match(name or ""):
        raise ValueError(f"Invalid database identifier: {name}")
    return name


def qn(name):
    return connection.ops.quote_name(assert_identifier(name))


def now_text():
    return timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")


def today_text():
    return timezone.localdate().isoformat()


def dict_rows(sql, params=None):
    from accounts.encryption import decrypt
    from django.conf import settings
    secret_key = getattr(settings, "SECRET_KEY", "fallback-key")
    
    with connection.cursor() as cursor:
        cursor.execute(sql, params or [])
        columns = [column[0] for column in cursor.description] if cursor.description else []
        rows = []
        for row in cursor.fetchall():
            row_dict = dict(zip(columns, row))
            for key, val in row_dict.items():
                if val and isinstance(val, str) and val.startswith("enc:"):
                    row_dict[key] = decrypt(val, secret_key)
            rows.append(row_dict)
        return rows


def one_row(sql, params=None):
    rows = dict_rows(sql, params)
    return rows[0] if rows else None


def table_exists(table_name):
    assert_identifier(table_name)
    with connection.cursor() as cursor:
        existing = connection.introspection.table_names(cursor)
    return table_name in existing


def table_columns(table_name):
    if not table_exists(table_name):
        return []
    with connection.cursor() as cursor:
        description = connection.introspection.get_table_description(cursor, table_name)
    return [getattr(column, "name", column[0]) for column in description]


def existing_columns(table_name, columns):
    available = set(table_columns(table_name))
    return [column for column in columns if column in available]


def primary_key_for(table_name, fallback=None):
    return fallback or DEFAULT_PRIMARY_KEYS.get(table_name)


def legacy_user_id(request):
    profile = getattr(request.user, "profile", None)
    return getattr(profile, "legacy_user_id", None) or None


def count_table(table_name):
    if not table_exists(table_name):
        return 0
    row = one_row(f"SELECT COUNT(*) AS total FROM {table_name}")
    return int(row["total"] or 0)


def current_user_role(user):
    profile = getattr(user, "profile", None)
    if profile and profile.role:
        return profile.role
    if user.is_superuser:
        return "Super Admin"
    return ""


def school_settings():
    if not table_exists("school_settings"):
        return {}
    settings = one_row("SELECT * FROM school_settings WHERE setting_id = 1") or {}
    try:
        from academic_structure.services import current_calendar

        snapshot = current_calendar()
        settings["current_term"] = snapshot.display_term
        settings["current_year"] = int(snapshot.display_year)
        settings["calendar_status"] = snapshot.status
        settings["next_term"] = snapshot.next_term.name if snapshot.next_term else ""
        settings["next_term_start_date"] = snapshot.next_term.start_date if snapshot.next_term else None
    except Exception:
        pass
    return settings


def audit_action(request, action, details=""):
    if not table_exists("audit_log"):
        return
    columns = table_columns("audit_log")
    data = {
        "user_id": legacy_user_id(request),
        "action": action,
        "details": details,
        "created_at": now_text(),
        "username": getattr(request.user, "username", ""),
        "user_role": current_user_role(request.user),
        "ip_address": request.META.get("REMOTE_ADDR", ""),
        "user_agent": request.META.get("HTTP_USER_AGENT", "")[:255],
        "path": request.path,
        "request_method": request.method,
    }
    usable = {key: value for key, value in data.items() if key in columns}
    if not usable:
        return
    names = list(usable)
    placeholders = ", ".join(["%s"] * len(names))
    sql = f"INSERT INTO {qn('audit_log')} ({', '.join(qn(name) for name in names)}) VALUES ({placeholders})"
    with connection.cursor() as cursor:
        cursor.execute(sql, [usable[name] for name in names])


def module_context(request, title, subtitle="", rows=None, columns=None, actions=None, stats=None):
    return {
        "title": title,
        "subtitle": subtitle,
        "rows": rows or [],
        "columns": columns or [],
        "actions": actions or [],
        "stats": stats or [],
        "settings": school_settings(),
    }


def export_filename(title, suffix):
    safe = re.sub(r"[^A-Za-z0-9_-]+", "-", title.lower()).strip("-") or "export"
    return f"{safe}.{suffix}"


def export_rows(title, rows, columns, export_type):
    from decimal import Decimal
    from datetime import date, datetime

    rows = list(rows or [])
    # Hydrate admission numbers if pupil_id is present in columns list
    has_pupil_id = "pupil_id" in columns
    if has_pupil_id:
        rows = hydrate_admission_numbers(rows)
        columns = [("admission_no" if col == "pupil_id" else col) for col in columns]

    if export_type == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{export_filename(title, "csv")}"'
        writer = csv.writer(response)
        writer.writerow(columns)
        for row in rows:
            writer.writerow([row.get(column, "") for column in columns])
        return response

    if export_type == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title[:31] or "Export"

        # Show Grid Lines
        sheet.views.sheetView[0].showGridLines = True

        # Styles
        title_font = Font(name="Segoe UI", size=16, bold=True, color="102A43")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")
        
        data_font = Font(name="Segoe UI", size=10)
        
        thin_side = Side(border_style="thin", color="D9E2EC")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Write Title block
        sheet.cell(row=1, column=1, value=title).font = title_font
        sheet.row_dimensions[1].height = 30
        sheet.append([])  # Empty row

        # Write Headers
        headers = [column_label(c) for c in columns]
        sheet.append(headers)
        header_row_idx = 3
        sheet.row_dimensions[header_row_idx].height = 25
        
        for col_idx in range(1, len(columns) + 1):
            cell = sheet.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Write Data
        for row in rows:
            row_data = []
            for col in columns:
                val = row.get(col, "")
                if val is None:
                    val = ""
                elif isinstance(val, (int, float, Decimal)) and not isinstance(val, bool):
                    val = float(val)
                row_data.append(val)
            sheet.append(row_data)

        # Style data rows
        start_row = 4
        end_row = start_row + len(rows) - 1
        for r_idx in range(start_row, end_row + 1):
            sheet.row_dimensions[r_idx].height = 20
            for col_idx in range(1, len(columns) + 1):
                cell = sheet.cell(row=r_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                val = cell.value
                col_name = columns[col_idx - 1].lower()
                
                # Align and format based on data type and column name
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if any(kw in col_name for kw in ["amount", "fee", "paid", "balance", "arrears", "credit", "debit", "money_in", "money_out"]):
                        cell.number_format = "$#,##0.00"
                    else:
                        cell.number_format = "#,##0"
                elif isinstance(val, (date, datetime)):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = "yyyy-mm-dd"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < 3:
                    continue
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{export_filename(title, "xlsx")}"'
        return response

    if export_type == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import mm

        # Landscape if > 6 columns
        if len(columns) > 6:
            pagesize = landscape(A4)
            width, height = pagesize
        else:
            pagesize = A4
            width, height = pagesize

        buffer = BytesIO()
        margin = 12 * mm
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=margin,
            leftMargin=margin,
            topMargin=margin,
            bottomMargin=margin
        )
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Title'],
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#102a43'),
            spaceAfter=15,
            alignment=0
        )
        
        header_cell_style = ParagraphStyle(
            'HeaderCell',
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.white,
            alignment=1
        )
        
        data_cell_style = ParagraphStyle(
            'DataCell',
            fontName='Helvetica',
            fontSize=8,
            textColor=colors.HexColor('#334e68'),
            leading=10
        )
        
        data_cell_right_style = ParagraphStyle(
            'DataCellRight',
            parent=data_cell_style,
            alignment=2
        )

        story = []
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 10))
        
        table_data = []
        # Header Row
        header_row = [Paragraph(column_label(col), header_cell_style) for col in columns]
        table_data.append(header_row)
        
        # Data Rows
        for row in rows:
            data_row = []
            for col in columns:
                val = row.get(col, "")
                if val is None:
                    val = ""
                
                col_name = col.lower()
                is_num = any(kw in col_name for kw in ["amount", "fee", "paid", "balance", "arrears", "credit", "debit", "money_in", "money_out"])
                
                if is_num and not isinstance(val, bool):
                    try:
                        formatted = f"${float(val):,.2f}"
                    except (ValueError, TypeError):
                        formatted = str(val)
                    p_style = data_cell_right_style
                else:
                    formatted = str(val)
                    p_style = data_cell_style
                    
                data_row.append(Paragraph(formatted, p_style))
            table_data.append(data_row)
            
        usable_width = width - (2 * margin)
        
        # Auto column width scaling
        col_widths = []
        for col in columns:
            max_len = len(column_label(col))
            for row in rows:
                val = str(row.get(col, "") or "")
                if len(val) > max_len:
                    max_len = len(val)
            col_widths.append(max_len)
            
        total_units = sum(col_widths)
        column_widths = [max(30.0, (w / total_units) * usable_width) for w in col_widths]
        total_scaled = sum(column_widths)
        column_widths = [(w / total_scaled) * usable_width for w in column_widths]
        
        report_table = Table(table_data, colWidths=column_widths, repeatRows=1)
        
        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#102a43')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d9e2ec')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#102a43')),
        ])
        
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f0f4f8'))
            else:
                t_style.add('BACKGROUND', (0, i), (-1, i), colors.white)
            t_style.add('TOPPADDING', (0, i), (-1, i), 6)
            t_style.add('BOTTOMPADDING', (0, i), (-1, i), 6)
            
        report_table.setStyle(t_style)
        story.append(report_table)
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{export_filename(title, "pdf")}"'
        return response

    raise Http404("Unsupported export type")


def build_querystring(request, **updates):
    data = request.GET.copy()
    for key, value in updates.items():
        if value is None:
            data.pop(key, None)
        else:
            data[key] = value
    query = data.urlencode()
    return f"?{query}" if query else ""


def clamp_int(value, default, minimum, maximum):
    try:
        number = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, number))


def is_database_id_column(column, pk_column=None):
    if not column:
        return False
    lowered = column.lower()
    return lowered == "id" or lowered == (pk_column or "").lower() or lowered.endswith("_id")


def display_columns_for(table_name, columns, pk_column=None):
    display_columns = []
    for column in columns:
        if column == "pupil_id":
            if "admission_no" not in display_columns:
                display_columns.append("admission_no")
            continue
        if is_database_id_column(column, pk_column=pk_column):
            continue
        display_columns.append(column)
    return display_columns


def column_label(column):
    labels = {
        "admission_no": "Admission No",
        "student_photo": "Photo",
        "amount_paid": "Amount Paid",
        "amount_billed": "Fees Charged",
        "amount_allocated": "Allocated",
        "balance": "Balance",
        "credit_balance": "Credit",
        "current_balance": "Current Balance",
        "guardian_name": "Parent/Guardian",
        "class_label": "Class",
        "class_stream": "Stream",
        "payment_method": "Method",
        "payment_date": "Date",
        "receipt_no": "Receipt No",
    }
    return labels.get(column, column.replace("_", " ").title())


def hydrate_student_display_fields(rows):
    rows = list(rows or [])
    if not rows:
        return rows
    try:
        from students.services import ensure_student_photo, student_age_text, student_photo_url
    except Exception:
        return rows
    for row in rows:
        if "date_of_birth" in row:
            row["age"] = student_age_text(row.get("date_of_birth"))
        if row.get("pupil_id") or row.get("photo_path"):
            photo_path = ensure_student_photo(row) if row.get("pupil_id") else row.get("photo_path")
            row["photo_path"] = photo_path or row.get("photo_path")
            row["photo_url"] = student_photo_url(row.get("photo_path"))
            row["student_photo"] = row["photo_url"]
    return rows


def academic_grade_number(grade=None, grade_id=None):
    text = str(grade or grade_id or "").strip()
    lowered = text.lower()
    if "completed o" in lowered or ("o level" in lowered and "completed" in lowered):
        return 7
    if "completed a" in lowered or ("a level" in lowered and "completed" in lowered):
        return 8
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def academic_grade_label(grade=None, grade_id=None):
    number = academic_grade_number(grade, grade_id)
    if 1 <= (number or 0) <= 6:
        return f"Form {number}"
    if number == 7:
        return "Completed O Level"
    if number == 8:
        return "Completed A Level"
    return str(grade or "").strip()


def academic_level_label(grade=None, grade_id=None):
    number = academic_grade_number(grade, grade_id)
    if 1 <= (number or 0) <= 4 or number == 7:
        return "O Level"
    if 5 <= (number or 0) <= 6 or number == 8:
        return "A Level"
    return ""


def academic_grade_option(value):
    return {"value": value, "label": academic_grade_label(value) or value}


def _clean_class_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def class_stream_name(grade=None, stream=None, grade_id=None, class_name=None, grade_name=None):
    stream_text = _clean_class_text(class_name if class_name is not None else stream)
    if not stream_text:
        return ""

    number = academic_grade_number(grade_name if grade_name is not None else grade, grade_id)
    if number:
        patterns = [
            rf"(?i)^form\s*{number}\b\s*[-/]?\s*",
            rf"(?i)^grade\s*{number}\b\s*[-/]?\s*",
            rf"(?i)^{number}\s*[-/]?\s*",
        ]
        for pattern in patterns:
            trimmed = re.sub(pattern, "", stream_text).strip()
            if trimmed and trimmed != stream_text:
                return _clean_class_text(trimmed)
    return stream_text


def class_stream_candidates(grade=None, stream=None, grade_id=None, class_name=None, grade_name=None):
    raw_stream = _clean_class_text(class_name if class_name is not None else stream)
    stream_only = class_stream_name(
        grade=grade,
        stream=stream,
        grade_id=grade_id,
        class_name=class_name,
        grade_name=grade_name,
    )
    number = academic_grade_number(grade_name if grade_name is not None else grade, grade_id)
    values = [raw_stream, stream_only]
    if number and stream_only:
        values.extend([
            f"Form {number} {stream_only}",
            f"{number}{stream_only}",
            f"{number} {stream_only}",
        ])
    if number and raw_stream and raw_stream != stream_only:
        values.extend([
            f"Form {number} {raw_stream}",
            f"{number}{raw_stream}",
            f"{number} {raw_stream}",
        ])
    return sorted({value.upper() for value in values if value})


def class_grade_candidates(grade=None, grade_id=None, grade_name=None):
    number = academic_grade_number(grade_name if grade_name is not None else grade, grade_id)
    values = [grade, grade_name, academic_grade_label(grade_name if grade_name is not None else grade, grade_id)]
    if number:
        values.extend([str(number), f"Form {number}", f"Grade {number}"])
    return sorted({str(value).strip().upper() for value in values if str(value or "").strip()})


def class_membership_where(selected_class, grade_name="", table_alias=""):
    if not selected_class:
        return "1 = 0", []

    prefix = f"{table_alias}." if table_alias else ""
    class_id = selected_class.get("class_id")
    grade_id = selected_class.get("grade_id")
    class_name = selected_class.get("class_name")
    stream_values = class_stream_candidates(
        grade_id=grade_id,
        grade_name=grade_name,
        class_name=class_name,
    )
    grade_values = class_grade_candidates(grade_id=grade_id, grade_name=grade_name)

    clauses = []
    params = []
    if class_id not in (None, ""):
        clauses.append(f"{prefix}class_id = %s")
        params.append(class_id)
    if grade_id not in (None, "") and stream_values:
        class_placeholders = ", ".join(["%s"] * len(stream_values))
        clauses.append(
            f"({prefix}grade_id = %s AND UPPER(TRIM(COALESCE({prefix}class_stream, ''))) "
            f"IN ({class_placeholders}))"
        )
        params.extend([grade_id] + stream_values)
    if grade_values and stream_values:
        grade_placeholders = ", ".join(["%s"] * len(grade_values))
        class_placeholders = ", ".join(["%s"] * len(stream_values))
        clauses.append(
            f"(UPPER(TRIM(COALESCE({prefix}grade, ''))) IN ({grade_placeholders}) "
            f"AND UPPER(TRIM(COALESCE({prefix}class_stream, ''))) IN ({class_placeholders}))"
        )
        params.extend(grade_values + stream_values)

    if not clauses:
        return "1 = 0", []
    return "(" + " OR ".join(clauses) + ")", params


def active_pupils_for_class(selected_class, grade_name="", select_fields=None):
    if not selected_class:
        return []
    fields = select_fields or "pupil_id, admission_no, first_name, surname"
    membership_where, params = class_membership_where(selected_class, grade_name)
    rows = dict_rows(
        f"""
        SELECT DISTINCT {fields}
        FROM pupils
        WHERE status = 'Active' AND {membership_where}
        ORDER BY surname, first_name
        """,
        params,
    )
    return hydrate_admission_numbers(rows)


def resolve_legacy_class_record(grade=None, stream=None, academic_year=None, grade_id=None):
    number = academic_grade_number(grade, grade_id)
    wanted_streams = set(class_stream_candidates(grade=grade, stream=stream, grade_id=number))
    if not number or not wanted_streams:
        return None

    where = ["grade_id = %s"]
    params = [number]
    if academic_year not in (None, ""):
        where.append("academic_year = %s")
        params.append(academic_year)
    classes = dict_rows(
        f"SELECT class_id, class_name, grade_id, academic_year FROM classes WHERE {' AND '.join(where)} ORDER BY academic_year DESC, class_id",
        params,
    )
    for item in classes:
        item_streams = set(class_stream_candidates(
            grade_id=item.get("grade_id"),
            grade_name=academic_grade_label(grade_id=item.get("grade_id")),
            class_name=item.get("class_name"),
        ))
        if wanted_streams & item_streams:
            return item
    return None


def compact_class_label(grade=None, stream=None, grade_id=None, class_name=None, grade_name=None):
    stream_text = class_stream_name(
        grade=grade,
        stream=stream,
        grade_id=grade_id,
        class_name=class_name,
        grade_name=grade_name,
    )
    if stream_text and re.fullmatch(r"\d+\s*[A-Za-z]+", stream_text):
        return stream_text.replace(" ", "").upper()

    grade_source = grade_name if grade_name is not None else grade
    grade_text = str(grade_source or "").strip()
    number = academic_grade_number(grade_text, grade_id)
    if number in {7, 8}:
        return academic_grade_label(grade_text, grade_id)
    if number:
        grade_part = f"Form {number}"
    elif grade_id not in {None, ""}:
        grade_part = str(grade_id).strip()
    else:
        grade_part = grade_text.replace("Grade", "").strip()

    if stream_text.lower() in {"", "all", "all streams", "none"}:
        return grade_part or stream_text
    return f"{grade_part} {stream_text}".strip()


def hydrate_class_labels(rows):
    rows = list(rows or [])
    grade_ids = sorted({row.get("grade_id") for row in rows if row.get("grade_id")})
    grade_names = {}
    if grade_ids and table_exists("grades"):
        placeholders = ", ".join(["%s"] * len(grade_ids))
        grade_names = {
            row["grade_id"]: row["grade_name"]
            for row in dict_rows(
                f"SELECT grade_id, grade_name FROM grades WHERE grade_id IN ({placeholders})",
                grade_ids,
            )
        }
    for row in rows:
        if row.get("class_label"):
            continue
        if "grade" in row:
            row["raw_grade"] = row.get("grade")
            row["academic_level"] = academic_level_label(row.get("grade"), row.get("grade_id"))
            row["grade"] = academic_grade_label(row.get("grade"), row.get("grade_id")) or row.get("grade")
        if "grade" in row or "class_stream" in row:
            row["class_label"] = compact_class_label(
                grade=row.get("grade"),
                stream=row.get("class_stream"),
                grade_id=row.get("grade_id"),
                grade_name=row.get("grade_name"),
            )
        elif "class_name" in row or "grade_id" in row or "grade_name" in row:
            row["grade_name"] = row.get("grade_name") or grade_names.get(row.get("grade_id"), "")
            row["class_label"] = compact_class_label(
                grade_id=row.get("grade_id"),
                grade_name=row.get("grade_name"),
                class_name=row.get("class_name"),
            )
    return rows


def hydrate_admission_numbers(rows):
    pupil_ids = sorted({row.get("pupil_id") for row in rows if row.get("pupil_id")})
    if not pupil_ids or not table_exists("pupils"):
        return rows
    placeholders = ", ".join(["%s"] * len(pupil_ids))
    pupils = dict_rows(
        f"SELECT pupil_id, admission_no FROM pupils WHERE pupil_id IN ({placeholders})",
        pupil_ids,
    )
    admissions = {row["pupil_id"]: row["admission_no"] for row in pupils}
    for row in rows:
        if row.get("pupil_id") in admissions:
            row.setdefault("admission_no", admissions[row["pupil_id"]])
    return rows


def pagination_context(request, total, page, per_page):
    total_pages = max(1, math.ceil(total / per_page)) if per_page else 1
    page = max(1, min(page, total_pages))
    return {
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
        "has_previous": page > 1,
        "has_next": page < total_pages,
        "previous_url": build_querystring(request, page=page - 1) if page > 1 else "",
        "next_url": build_querystring(request, page=page + 1) if page < total_pages else "",
    }


def render_rows_page(
    request,
    title,
    rows,
    columns,
    subtitle="",
    actions=None,
    row_actions=None,
    stats=None,
    total=None,
    page=None,
    per_page=None,
    filters=None,
    extra_context=None,
):
    rows = hydrate_admission_numbers(list(rows or []))
    rows = hydrate_class_labels(rows)
    if any(column in {"student_photo", "age"} for column in columns):
        rows = hydrate_student_display_fields(rows)
    rows = prepare_row_actions(rows, row_actions)
    q = (request.GET.get("q") or "").strip()
    page = page or clamp_int(request.GET.get("page"), 1, 1, 1000000)
    per_page = per_page or clamp_int(request.GET.get("per_page"), 25, 10, 100)
    total = len(rows) if total is None else int(total or 0)
    context = module_context(
        request,
        title,
        subtitle,
        rows=rows,
        columns=columns,
        actions=actions or [],
        stats=stats or [("Records", total), ("Search", q or "All")],
    )
    context.update(
        {
            "q": q,
            "filters": filters or [],
            "has_row_actions": bool(row_actions),
            "column_labels": {column: column_label(column) for column in columns},
            "pagination": pagination_context(request, total, page, per_page),
            "per_page_options": [10, 25, 50, 100],
        }
    )
    if extra_context:
        context.update(extra_context)
    return render(request, "school/table_page.html", context)


def format_action_href(pattern, row):
    if not pattern:
        return "#"
    try:
        return pattern.format(**row)
    except KeyError:
        return "#"


def prepare_row_actions(rows, row_actions):
    if not row_actions:
        return rows
    prepared = []
    for row in rows:
        item = dict(row)
        actions = []
        for action in row_actions:
            action_copy = dict(action)
            href = format_action_href(action.get("href"), item)
            if href == "#":
                continue
            action_copy["href"] = href
            actions.append(action_copy)
        item["row_actions"] = actions
        prepared.append(item)
    return prepared


def render_table_page(
    request,
    title,
    table_name,
    columns,
    subtitle="",
    order_by=None,
    search_columns=None,
    filters=None,
    where=None,
    params=None,
    actions=None,
    row_actions=None,
    pk_column=None,
    create_href=None,
    create_label=None,
    extra_context=None,
):
    rows = []
    total = 0
    q = (request.GET.get("q") or "").strip()
    page = clamp_int(request.GET.get("page"), 1, 1, 1000000)
    per_page = clamp_int(request.GET.get("per_page"), 25, 10, 100)
    filter_context = []
    requested_columns = list(columns)
    wants_class_label = "class_label" in requested_columns
    virtual_columns = {"student_photo", "age", "academic_level"} if table_name == "pupils" else set()
    requested_virtual_columns = [column for column in requested_columns if column in virtual_columns]
    display_columns = list(requested_columns)
    clauses = []
    query_params = list(params or [])
    if table_exists(table_name):
        table_cols = table_columns(table_name)
        db_requested_columns = [column for column in requested_columns if column not in {"class_label", *virtual_columns}]
        columns = existing_columns(table_name, db_requested_columns)
        if not columns:
            columns = table_cols
        pk_column = primary_key_for(table_name, pk_column)
        query_columns = list(columns)
        if pk_column and pk_column in table_cols and pk_column not in query_columns:
            query_columns.append(pk_column)
        if wants_class_label:
            for source_column in ["grade", "class_stream", "grade_id", "class_id", "class_name", "grade_name"]:
                if source_column in table_cols and source_column not in query_columns:
                    query_columns.append(source_column)
            display_columns = []
            for column in requested_columns:
                if column == "class_label":
                    display_columns.append(column)
                elif column in columns and not is_database_id_column(column, pk_column=pk_column):
                    display_columns.append(column)
        else:
            display_columns = display_columns_for(table_name, columns, pk_column=pk_column)
        if requested_virtual_columns:
            display_columns = []
            for column in requested_columns:
                if column in requested_virtual_columns:
                    display_columns.append(column)
                elif column == "class_label":
                    display_columns.append(column)
                elif column in columns and not is_database_id_column(column, pk_column=pk_column):
                    display_columns.append(column)
            for source_column in ["pupil_id", "photo_path", "date_of_birth", "admission_no", "grade", "grade_id"]:
                if source_column in table_cols and source_column not in query_columns:
                    query_columns.append(source_column)
        if "pupil_id" in query_columns and "admission_no" not in query_columns and table_name == "pupils":
            query_columns.append("admission_no")
        sql = f"SELECT {', '.join(qn(column) for column in query_columns)} FROM {qn(table_name)}"
        count_sql = f"SELECT COUNT(*) AS total FROM {qn(table_name)}"
        if where:
            clauses.append(f"({where})")
        if q and search_columns:
            searchable = existing_columns(table_name, search_columns)
            if searchable:
                clauses.append("(" + " OR ".join(f"CAST({qn(column)} AS TEXT) LIKE %s" for column in searchable) + ")")
                query_params.extend([f"%{q}%"] * len(searchable))
        for filter_def in filters or []:
            name = filter_def.get("name")
            if name not in table_columns(table_name):
                continue
            value = (request.GET.get(name) or "").strip()
            item = dict(filter_def)
            item.setdefault("label", column_label(name))
            item["value"] = value
            if "options" not in item:
                options = [
                    row[name]
                    for row in dict_rows(
                        f"SELECT DISTINCT {qn(name)} FROM {qn(table_name)} WHERE {qn(name)} IS NOT NULL AND TRIM(CAST({qn(name)} AS TEXT)) != '' ORDER BY {qn(name)} LIMIT 100"
                    )
                ]
                item["options"] = [academic_grade_option(value) for value in options] if table_name == "pupils" and name == "grade" else options
            filter_context.append(item)
            if value:
                clauses.append(f"CAST({qn(name)} AS TEXT) = %s")
                query_params.append(value)
        if clauses:
            where_sql = " WHERE " + " AND ".join(clauses)
            sql += where_sql
            count_sql += where_sql
        total_row = one_row(count_sql, query_params)
        total = int(total_row["total"] or 0) if total_row else 0
        page = min(page, max(1, math.ceil(total / per_page))) if total else 1
        if order_by:
            sql += f" ORDER BY {order_by}"
        sql += " LIMIT %s OFFSET %s"
        rows = dict_rows(sql, query_params + [per_page, (page - 1) * per_page])
        rows = hydrate_admission_numbers(rows)
        rows = hydrate_class_labels(rows)
        if requested_virtual_columns:
            rows = hydrate_student_display_fields(rows)
    export_type = (request.GET.get("export") or "").lower()
    if export_type:
        export_rows_data = rows
        if table_exists(table_name):
            export_sql = f"SELECT {', '.join(qn(column) for column in query_columns)} FROM {qn(table_name)}"
            if clauses:
                export_sql += " WHERE " + " AND ".join(clauses)
            if order_by:
                export_sql += f" ORDER BY {order_by}"
            export_sql += " LIMIT 5000"
            export_rows_data = hydrate_class_labels(hydrate_admission_numbers(dict_rows(export_sql, query_params)))
            if requested_virtual_columns:
                export_rows_data = hydrate_student_display_fields(export_rows_data)
        return export_rows(title, export_rows_data, display_columns, export_type)
    top_actions = list(actions or [])
    if create_href:
        top_actions.insert(0, {"label": create_label or "Add Record", "href": create_href, "icon": "bi-plus-circle"})
    for export_type, icon in [("csv", "bi-filetype-csv"), ("xlsx", "bi-file-earmark-excel"), ("pdf", "bi-file-earmark-pdf")]:
        top_actions.append(
            {
                "label": export_type.upper(),
                "href": build_querystring(request, export=export_type),
                "icon": icon,
            }
        )
    if pk_column and row_actions is None:
        row_actions = [
            {"label": "View", "href": f"{request.path.rstrip('/')}/{{{pk_column}}}", "icon": "bi-eye", "class": "btn-outline-primary"},
            {"label": "Edit", "href": f"{request.path.rstrip('/')}/{{{pk_column}}}/edit", "icon": "bi-pencil", "class": "btn-outline-secondary"},
            {
                "label": "Delete",
                "href": f"{request.path.rstrip('/')}/{{{pk_column}}}/delete",
                "icon": "bi-trash",
                "class": "btn-outline-danger",
                "method": "post",
                "confirm": "Delete this record?",
            },
        ]
    rows = prepare_row_actions(rows, row_actions)
    context = module_context(
        request,
        title,
        subtitle,
        rows=rows,
        columns=display_columns,
        actions=top_actions,
        stats=[("Records", total), ("Search", q or "All")],
    )
    context["q"] = q
    context["table_name"] = table_name
    context["filters"] = filter_context
    context["has_row_actions"] = bool(row_actions)
    context["column_labels"] = {column: column_label(column) for column in display_columns}
    context["pagination"] = pagination_context(request, total, page, per_page)
    context["per_page_options"] = [10, 25, 50, 100]
    if extra_context:
        context.update(extra_context)
    return render(request, "school/table_page.html", context)


def render_detail_page(request, title, table_name, pk_column, pk_value, actions=None, pdf_filename=None):
    row = None
    if table_exists(table_name):
        row = one_row(f"SELECT * FROM {qn(table_name)} WHERE {qn(pk_column)} = %s", [pk_value])
    if row is None:
        messages.error(request, f"{title} was not found.")
        return redirect("accounts:dashboard")
    display_row = {}
    if row.get("pupil_id"):
        hydrated = hydrate_admission_numbers([dict(row)])
        if hydrated and hydrated[0].get("admission_no"):
            display_row["admission_no"] = hydrated[0]["admission_no"]
    for key, value in row.items():
        if key == "pupil_id" or is_database_id_column(key, pk_column=pk_column):
            continue
        display_row[key] = value
    if request.GET.get("format") == "pdf":
        lines = [f"{key}: {value or '-'}" for key, value in display_row.items()]
        return simple_pdf(title, lines, pdf_filename or export_filename(title, "pdf"))
    return render(
        request,
        "school/detail_page.html",
        {"title": title, "row": display_row, "actions": actions or [], "settings": school_settings()},
    )


def simple_form_page(request, title, subtitle="", fields=None):
    if request.method == "POST":
        messages.info(request, "This page is now served by Django. Detailed save logic is being migrated module by module.")
    return render(
        request,
        "school/form_page.html",
        {"title": title, "subtitle": subtitle, "fields": fields or [], "settings": school_settings()},
    )


def normalize_form_fields(table_name, fields=None, pk_column=None, values=None):
    columns = table_columns(table_name)
    if fields is None:
        names = [column for column in columns if column != pk_column and column not in READ_ONLY_COLUMNS]
        fields = [{"name": name, "label": name.replace("_", " ").title()} for name in names]
    normalized = []
    for field in fields:
        if isinstance(field, str):
            field = {"name": field, "label": field.replace("_", " ").title()}
        name = field.get("name")
        if not name or name not in columns or name == pk_column:
            continue
        item = dict(field)
        item.setdefault("label", name.replace("_", " ").title())
        if name in {"notes", "address", "remarks", "details", "description", "message_body", "instructions", "teacher_comment", "provider_response"}:
            item.setdefault("widget", "textarea")
        elif name == "role":
            item.setdefault("widget", "select")
            item.setdefault(
                "options",
                [
                    "Super Admin",
                    "Administrator",
                    "Headmaster",
                    "Headmaster / Headmistress",
                    "Deputy Head",
                    "HOD",
                    "Bursar / Accounts Clerk",
                    "Accountant",
                    "Registrar / Office Clerk",
                    "Clerk",
                    "Teacher",
                    "Librarian",
                    "Transport Staff",
                    "Hostel Staff",
                    "Nurse",
                    "Parent",
                    "Student",
                ],
            )
        elif name == "status":
            item.setdefault("widget", "select")
            item.setdefault("options", ["Active", "Inactive", "Draft", "Published", "Paid", "Pending", "Approved", "Rejected", "Returned"])
        elif name in {"school_logo", "school_stamp"} or name.endswith("_logo") or name.endswith("_stamp"):
            item.setdefault("type", "file")
        elif "date" in name or name.endswith("_on"):
            item.setdefault("type", "date")
        elif any(token in name for token in ["amount", "salary", "mark", "score", "total", "average", "quantity", "year", "period_no"]):
            item.setdefault("type", "number")
        else:
            item.setdefault("type", "text")
        if values is not None:
            item["value"] = values.get(name)
        normalized.append(item)
    return normalized


def clean_post_value(value):
    if value is None:
        return None
    value = value.strip()
    return value if value != "" else None


def add_system_defaults(request, table_name, data, update=False):
    columns = table_columns(table_name)
    for column in DATE_DEFAULT_COLUMNS:
        if column in columns and column not in data and not update:
            data[column] = today_text()
    if "created_at" in columns and "created_at" not in data and not update:
        data["created_at"] = now_text()
    if "updated_at" in columns:
        data["updated_at"] = now_text()
    if "marked_at" in columns and "marked_at" not in data and not update:
        data["marked_at"] = now_text()
    if "generated_at" in columns and "generated_at" not in data and not update:
        data["generated_at"] = now_text()
    if "issued_date" in columns and "issued_date" not in data and not update:
        data["issued_date"] = today_text()
    user_id = legacy_user_id(request)
    if user_id:
        for column in SYSTEM_USER_COLUMNS:
            if column in columns and column not in data:
                data[column] = user_id
    return data


def insert_record(request, table_name, data):
    from accounts.encryption import encrypt
    from django.conf import settings
    secret_key = getattr(settings, "SECRET_KEY", "fallback-key")
    
    columns = table_columns(table_name)
    usable = {}
    for key, value in data.items():
        if key in columns:
            if key in {"guardian_phone", "guardian_email", "guardian_name", "date_of_birth", "phone_number"} and value:
                val_str = str(value)
                if not val_str.startswith("enc:"):
                    usable[key] = encrypt(val_str, secret_key)
                else:
                    usable[key] = val_str
            else:
                usable[key] = value
                
    if not usable:
        raise ValueError("No valid fields were submitted.")
    names = list(usable)
    placeholders = ", ".join(["%s"] * len(names))
    
    pk_col = primary_key_for(table_name)
    is_postgres = (connection.vendor == 'postgresql')
    if is_postgres and pk_col:
        sql = f"INSERT INTO {qn(table_name)} ({', '.join(qn(name) for name in names)}) VALUES ({placeholders}) RETURNING {qn(pk_col)}"
    else:
        sql = f"INSERT INTO {qn(table_name)} ({', '.join(qn(name) for name in names)}) VALUES ({placeholders})"
        
    with connection.cursor() as cursor:
        cursor.execute(sql, [usable[name] for name in names])
        if is_postgres and pk_col:
            res = cursor.fetchone()
            return res[0] if res else None
        return getattr(cursor, "lastrowid", None)


def update_record(request, table_name, pk_column, pk_value, data):
    from accounts.encryption import encrypt
    from django.conf import settings
    secret_key = getattr(settings, "SECRET_KEY", "fallback-key")
    
    columns = table_columns(table_name)
    usable = {}
    for key, value in data.items():
        if key in columns and key != pk_column:
            if key in {"guardian_phone", "guardian_email", "guardian_name", "date_of_birth", "phone_number"} and value:
                val_str = str(value)
                if not val_str.startswith("enc:"):
                    usable[key] = encrypt(val_str, secret_key)
                else:
                    usable[key] = val_str
            else:
                usable[key] = value
                
    if not usable:
        raise ValueError("No valid fields were submitted.")
    assignments = ", ".join(f"{qn(name)} = %s" for name in usable)
    sql = f"UPDATE {qn(table_name)} SET {assignments} WHERE {qn(pk_column)} = %s"
    with connection.cursor() as cursor:
        cursor.execute(sql, [usable[name] for name in usable] + [pk_value])


def render_record_form_page(
    request,
    title,
    table_name,
    fields=None,
    pk_column=None,
    pk_value=None,
    subtitle="",
    redirect_to=None,
    extra_defaults=None,
):
    pk_column = primary_key_for(table_name, pk_column)
    row = {}
    if pk_value is not None:
        row = one_row(f"SELECT * FROM {qn(table_name)} WHERE {qn(pk_column)} = %s", [pk_value])
        if row is None:
            messages.error(request, f"{title} was not found.")
            return redirect(redirect_to or "accounts:dashboard")
    form_fields = normalize_form_fields(table_name, fields, pk_column=pk_column, values=row)
    if request.method == "POST":
        data = {}
        for field in form_fields:
            name = field["name"]
            if field.get("type") == "file":
                if name in request.FILES:
                    import os
                    from django.conf import settings as django_settings
                    uploaded_file = request.FILES[name]
                    static_dir = os.path.join(django_settings.BASE_DIR, "static", "img")
                    os.makedirs(static_dir, exist_ok=True)
                    file_name = uploaded_file.name
                    dest_path = os.path.join(static_dir, file_name)
                    with open(dest_path, "wb") as f:
                        for chunk in uploaded_file.chunks():
                            f.write(chunk)
                    data[name] = f"img/{file_name}"
                else:
                    data[name] = row.get(name) if (pk_value is not None and row) else None
            else:
                data[name] = clean_post_value(request.POST.get(name))
        data.update(extra_defaults or {})
        data = add_system_defaults(request, table_name, data, update=pk_value is not None)
        try:
            if pk_value is None:
                new_id = insert_record(request, table_name, data)
                audit_action(request, f"Create {table_name}", f"Created {table_name} record {new_id or ''}".strip())
                messages.success(request, f"{title} saved.")
            else:
                update_record(request, table_name, pk_column, pk_value, data)
                audit_action(request, f"Update {table_name}", f"Updated {table_name} record {pk_value}")
                messages.success(request, f"{title} updated.")
            return redirect(redirect_to or request.path.rsplit("/", 2)[0] or "accounts:dashboard")
        except Exception as exc:
            messages.error(request, f"Could not save {title}: {exc}")
            for field in form_fields:
                field["value"] = request.POST.get(field["name"], "")
    return render(
        request,
        "school/form_page.html",
        {"title": title, "subtitle": subtitle, "fields": form_fields, "settings": school_settings()},
    )


def delete_record(request, title, table_name, pk_column, pk_value, redirect_to):
    row = one_row(f"SELECT * FROM {qn(table_name)} WHERE {qn(pk_column)} = %s", [pk_value]) if table_exists(table_name) else None
    if row is None:
        messages.error(request, f"{title} was not found.")
        return redirect(redirect_to)
    if request.method == "POST":
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"DELETE FROM {qn(table_name)} WHERE {qn(pk_column)} = %s", [pk_value])
            audit_action(request, f"Delete {table_name}", f"Deleted {table_name} record {pk_value}")
            messages.success(request, f"{title} deleted.")
        except Exception as exc:
            messages.error(request, f"Could not delete {title}: {exc}")
        return redirect(redirect_to)
    display_row = {}
    if row.get("pupil_id"):
        hydrated = hydrate_admission_numbers([dict(row)])
        if hydrated and hydrated[0].get("admission_no"):
            display_row["admission_no"] = hydrated[0]["admission_no"]
    for key, value in row.items():
        if key == "pupil_id" or is_database_id_column(key, pk_column=pk_column):
            continue
        display_row[key] = value
    return render(
        request,
        "school/detail_page.html",
        {
            "title": f"Delete {title}",
            "row": display_row,
            "settings": school_settings(),
            "delete_confirm": True,
            "actions": [{"label": "Cancel", "href": redirect_to, "icon": "bi-x-circle"}],
        },
    )


def update_record_fields(request, table_name, pk_column, pk_value, fields, success_message, redirect_to):
    try:
        data = add_system_defaults(request, table_name, dict(fields), update=True)
        update_record(request, table_name, pk_column, pk_value, data)
        audit_action(request, f"Update {table_name}", success_message)
        messages.success(request, success_message)
    except Exception as exc:
        messages.error(request, f"Could not update record: {exc}")
    return redirect(redirect_to)


def get_pdf_header(settings, width_pts):
    import os
    from django.conf import settings as django_settings
    from reportlab.platypus import Table, TableStyle, Image, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    logo_path = None
    custom_logo = settings.get("school_logo")
    if custom_logo:
        test_path = os.path.join(django_settings.BASE_DIR, "static", custom_logo)
        if os.path.exists(test_path) and not test_path.lower().endswith(".svg"):
            logo_path = test_path
        else:
            png_test = test_path.rsplit(".", 1)[0] + ".png"
            if os.path.exists(png_test):
                logo_path = png_test
                
    if not logo_path or not os.path.exists(logo_path):
        logo_path = os.path.join(django_settings.BASE_DIR, "static", "img", "raydon-system-logo.png")
    
    styles = getSampleStyleSheet()
    details_style = ParagraphStyle(
        'HeaderSchoolDetails',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#334155'),
        leading=12
    )
    
    details_text = f"<font size=\"14\" color=\"#0f766e\"><b>{settings.get('school_name', 'RAYDON HIGH SCHOOL')}</b></font>"
    if settings.get("school_motto"):
        details_text += f"<br/><font size=\"8.5\" color=\"#64748b\"><i>\"{settings.get('school_motto')}\"</i></font>"
    details_text += "<br/>"
    
    details_parts = []
    if settings.get("school_address"):
        details_parts.append(settings.get("school_address"))
    phone = settings.get("school_phone")
    email = settings.get("school_email")
    website = settings.get("school_website")
    contact_parts = []
    if phone:
        contact_parts.append(f"Phone: {phone}")
    if email:
        contact_parts.append(f"Email: {email}")
    if website:
        contact_parts.append(f"Web: {website}")
    if contact_parts:
        details_parts.append(" | ".join(contact_parts))
    
    details_text += "<br/>".join(details_parts)
    
    logo_flowable = None
    if logo_path and os.path.exists(logo_path):
        try:
            logo_flowable = Image(logo_path, width=50, height=50)
        except Exception as e:
            pass
            
    if logo_flowable:
        data = [[logo_flowable, Paragraph(details_text, details_style)]]
        header_table = Table(data, colWidths=[60, width_pts - 60])
    else:
        data = [[Paragraph(details_text, details_style)]]
        header_table = Table(data, colWidths=[width_pts])
        
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (1, 0), (1, 0), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
    ]))
    return header_table


def simple_pdf(title, lines, filename):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    for line in lines:
        story.append(Paragraph(str(line), styles["Normal"]))
        story.append(Spacer(1, 6))
    document.build(story)
    buffer.seek(0)
    return FileResponse(buffer, content_type="application/pdf", as_attachment=True, filename=filename)


def date_label(value):
    return str(value or date.today().isoformat())


def json_success(payload=None):
    return JsonResponse(payload or {"status": "ok"})
