from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.db import connection, transaction
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from school_system_django.native import school_settings
from academics.views import active_academic_year
from accounts.permissions import permission_required, normalized_role, assigned_classes_for_teacher
from students.models import SchoolClass, Pupil
from academics.models import Subject, ClassTimetableEntry
from accounts.models import UserProfile
from timetable.models import Room, SubjectAllocation, TeacherAvailability, TimetablePeriodConfig
from timetable.forms import RoomForm, SubjectAllocationForm, TeacherAvailabilityForm, TimetablePeriodConfigForm
from timetable.services.scheduler import TimetableScheduler
from timetable.services.conflict_checker import ConflictChecker
from timetable.services.workload_manager import WorkloadManager
from timetable.services.pdf_generator import TimetablePDFGenerator

@permission_required("timetable.view")
def timetable_dashboard(request):
    role = normalized_role(request.user)
    if role == 'Teacher':
        allowed_class_ids = assigned_classes_for_teacher(request.user)
        if allowed_class_ids:
            return redirect(f"/timetables/grid?type=class&target={allowed_class_ids[0]}")
        else:
            from django.contrib import messages
            messages.error(request, "You are not assigned as a class teacher to any class.")
            return redirect("accounts:dashboard")
            
    current_year = active_academic_year()
    
    # 1. Basic Stats
    total_classes = SchoolClass.objects.filter(academic_year=current_year).count()
    total_teachers = UserProfile.objects.filter(role='Teacher', status='Active').count()
    
    # Count how many classes have at least one slot scheduled
    scheduled_classes_count = ClassTimetableEntry.objects.filter(
        academic_year=current_year
    ).values('class_id').distinct().count()
    
    # Scan for conflicts in memory (efficient)
    entries = list(ClassTimetableEntry.objects.filter(academic_year=current_year))
    conflicts = []
    teacher_slots = {}
    room_slots = {}
    class_slots = {}
    
    # Load class name mapping for display
    class_names = {c.class_id: c.class_name for c in SchoolClass.objects.filter(academic_year=current_year)}
    
    for e in entries:
        c_name = class_names.get(e.class_id, f"Class ID {e.class_id}")
        
        # Class check
        class_key = (e.class_id, e.day_name, e.period_no)
        if class_key in class_slots:
            other = class_slots[class_key]
            conflicts.append({
                "type": "Class Double Booking",
                "message": f"Class '{c_name}' has both '{e.subject_name}' and '{other.subject_name}' scheduled at {e.day_name} Period {e.period_no}."
            })
        else:
            class_slots[class_key] = e
            
        # Teacher check
        if e.teacher_name:
            t_name = e.teacher_name.upper().strip()
            t_key = (t_name, e.day_name, e.period_no)
            if t_key in teacher_slots:
                other = teacher_slots[t_key]
                other_class = class_names.get(other.class_id, f"Class ID {other.class_id}")
                conflicts.append({
                    "type": "Teacher Double Booking",
                    "message": f"Teacher '{e.teacher_name}' is assigned to teach both '{c_name}' and '{other_class}' at {e.day_name} Period {e.period_no}."
                })
            else:
                teacher_slots[t_key] = e
                
        # Room check
        if e.room_name:
            r_name = e.room_name.upper().strip()
            r_key = (r_name, e.day_name, e.period_no)
            if r_key in room_slots:
                other = room_slots[r_key]
                other_class = class_names.get(other.class_id, f"Class ID {other.class_id}")
                conflicts.append({
                    "type": "Room Double Booking",
                    "message": f"Room '{e.room_name}' is assigned to both '{c_name}' and '{other_class}' at {e.day_name} Period {e.period_no}."
                })
            else:
                room_slots[r_key] = e

    # Calculate free periods count
    lesson_periods_count = TimetablePeriodConfig.objects.filter(period_type='Lesson').count()
    if lesson_periods_count == 0:
        lesson_periods_count = 8 # Fallback
    
    total_teacher_slots = total_teachers * 5 * lesson_periods_count
    assigned_slots = len(teacher_slots)
    free_periods_count = max(0, total_teacher_slots - assigned_slots)

    # Workloads summary
    workloads = WorkloadManager.get_all_workloads_summary(current_year)

    context = {
        "current_year": current_year,
        "total_classes": total_classes,
        "total_teachers": total_teachers,
        "total_timetables": scheduled_classes_count,
        "conflicts_count": len(conflicts),
        "conflicts": conflicts[:10], # Show first 10 conflicts
        "free_periods": free_periods_count,
        "workloads": workloads,
        "settings": school_settings()
    }
    return render(request, "timetable/dashboard.html", context)


def get_grid_data(view_type, target, academic_year):
    """
    Helper function to build rows and cells for Class, Teacher, Room, and Subject grids.
    """
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    periods = list(TimetablePeriodConfig.objects.all().order_by('period_no'))
    if not periods:
        TimetableScheduler.pre_populate_configs_if_empty()
        periods = list(TimetablePeriodConfig.objects.all().order_by('period_no'))

    # Load matching timetable entries
    query = ClassTimetableEntry.objects.filter(academic_year=academic_year)
    if view_type == 'class' and target:
        query = query.filter(class_id=target)
    elif view_type == 'teacher' and target:
        query = query.filter(teacher_name__iexact=target.strip())
    elif view_type == 'room' and target:
        query = query.filter(room_name__iexact=target.strip())
    elif view_type == 'subject' and target:
        query = query.filter(subject_id=target)
        
    entries = list(query)
    
    # Class mapping for extra labels
    classes_dict = {c.class_id: c.class_name for c in SchoolClass.objects.all()}

    # Format into grid
    grid_rows = []
    for day in days:
        row_cells = []
        for p in periods:
            # If period is Break/Lunch, cell is empty (labeled via config)
            if p.period_type in ['Break', 'Lunch']:
                row_cells.append({
                    "period_type": p.period_type,
                    "label": p.label or p.period_type,
                    "is_break": True
                })
                continue
                
            # Find matching entry
            match = None
            for e in entries:
                if e.day_name == day and e.period_no == p.period_no:
                    match = e
                    break
            
            if match:
                # Add class label if view is teacher/room/subject
                extra_label = ""
                if view_type != 'class':
                    extra_label = classes_dict.get(match.class_id, f"Class {match.class_id}")

                row_cells.append({
                    "timetable_id": match.timetable_id,
                    "class_id": match.class_id,
                    "class_name": classes_dict.get(match.class_id, ""),
                    "subject_id": match.subject_id,
                    "subject_name": match.subject_name,
                    "teacher_name": match.teacher_name or "-",
                    "room_name": match.room_name or "-",
                    "is_locked": match.is_locked,
                    "extra_label": extra_label,
                    "is_break": False
                })
            else:
                row_cells.append({
                    "is_empty": True,
                    "is_break": False,
                    "day_name": day,
                    "period_no": p.period_no
                })
        grid_rows.append({
            "day_name": day,
            "cells": row_cells
        })
        
    return grid_rows, periods


@permission_required("timetable.view")
def timetable_grid(request):
    current_year = active_academic_year()
    view_type = request.GET.get('type', 'class')
    target = request.GET.get('target', '')
    
    role = normalized_role(request.user)
    if role == 'Teacher':
        allowed_class_ids = assigned_classes_for_teacher(request.user)
        if not allowed_class_ids:
            from django.contrib import messages
            messages.error(request, "You are not assigned as a class teacher to any class.")
            return redirect("accounts:dashboard")
        view_type = 'class'
        if not target or int(target) not in allowed_class_ids:
            target = str(allowed_class_ids[0])
            
    # Dropdowns for selector
    if role == 'Teacher':
        classes = SchoolClass.objects.filter(academic_year=current_year, class_id__in=allowed_class_ids).order_by('class_name')
        teachers = UserProfile.objects.none()
        rooms = Room.objects.none()
        subjects = Subject.objects.none()
    else:
        classes = SchoolClass.objects.filter(academic_year=current_year).order_by('class_name')
        teachers = UserProfile.objects.filter(role='Teacher', status='Active').order_by('full_name')
        rooms = Room.objects.all().order_by('room_name')
        subjects = Subject.objects.filter(status='Active').order_by('subject_name')
    
    timetable_rows = []
    periods = []
    selected_name = ""
    
    # Check if a target has been loaded, or choose the first class if none specified
    if not target and view_type == 'class' and classes.exists():
        target = str(classes.first().class_id)
        
    if target:
        if view_type == 'class':
            class_obj = get_object_or_404(SchoolClass, class_id=target)
            selected_name = class_obj.class_name
            timetable_rows, periods = get_grid_data('class', int(target), current_year)
        elif view_type == 'teacher':
            selected_name = target
            timetable_rows, periods = get_grid_data('teacher', target, current_year)
        elif view_type == 'room':
            selected_name = target
            timetable_rows, periods = get_grid_data('room', target, current_year)
        elif view_type == 'subject':
            subject_obj = get_object_or_404(Subject, subject_id=target)
            selected_name = subject_obj.subject_name
            timetable_rows, periods = get_grid_data('subject', int(target), current_year)

    # Master Board (special view showing all classes for a selected day)
    selected_day = request.GET.get('day', 'Monday')
    master_grid = []
    if view_type == 'master':
        all_periods = list(TimetablePeriodConfig.objects.all().order_by('period_no'))
        for school_class in classes:
            class_cells = []
            for p in all_periods:
                if p.period_type in ['Break', 'Lunch']:
                    class_cells.append({
                        "period_type": p.period_type,
                        "label": p.label or p.period_type,
                        "is_break": True
                    })
                    continue
                entry = ClassTimetableEntry.objects.filter(
                    class_id=school_class.class_id,
                    day_name=selected_day,
                    period_no=p.period_no,
                    academic_year=current_year
                ).first()
                if entry:
                    class_cells.append({
                        "timetable_id": entry.timetable_id,
                        "subject_name": entry.subject_name,
                        "teacher_name": entry.teacher_name,
                        "room_name": entry.room_name,
                        "is_locked": entry.is_locked,
                        "is_break": False
                    })
                else:
                    class_cells.append({
                        "is_empty": True,
                        "is_break": False
                    })
            master_grid.append({
                "class_name": school_class.class_name,
                "cells": class_cells
            })
        periods = all_periods

    context = {
        "view_type": view_type,
        "target": target,
        "selected_name": selected_name,
        "selected_day": selected_day,
        "classes": classes,
        "teachers": teachers,
        "rooms": rooms,
        "subjects": subjects,
        "timetable_rows": timetable_rows,
        "periods": periods,
        "master_grid": master_grid,
        "settings": school_settings()
    }
    return render(request, "timetable/grid_view.html", context)


@permission_required("timetable.manage")
def timetable_generate(request):
    current_year = active_academic_year()
    classes = SchoolClass.objects.filter(academic_year=current_year).order_by('class_name')
    
    if request.method == "POST":
        class_id = request.POST.get('class_id')
        replace = request.POST.get('replace') == 'yes'
        target_id = int(class_id) if class_id else None
        
        result = TimetableScheduler.generate(current_year, target_class_id=target_id, replace_existing=replace)
        return JsonResponse({"status": "success", "message": f"Successfully generated {result['created']} entries. Skipped {result['skipped']} class(es)."})
        
    return render(request, "timetable/generate.html", {
        "classes": classes,
        "current_year": current_year,
        "settings": school_settings()
    })


@permission_required("timetable.manage")
def api_check_conflict(request):
    role = normalized_role(request.user)
    if role == 'Teacher':
        return JsonResponse({"status": "error", "message": "Permission denied"}, status=403)
    """
    POST API to check conflicts for a prospective slot assignment.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST method required"}, status=400)
        
    class_id = request.POST.get('class_id')
    teacher_name = request.POST.get('teacher_name', '')
    room_name = request.POST.get('room_name', '')
    day_name = request.POST.get('day_name')
    period_no = request.POST.get('period_no')
    exclude_id = request.POST.get('exclude_id')
    
    if not (class_id and day_name and period_no):
        return JsonResponse({"status": "error", "message": "Missing required fields"}, status=400)
        
    current_year = active_academic_year()
    exclude_id_int = int(exclude_id) if exclude_id else None
    
    conflicts = ConflictChecker.check_conflicts(
        class_id=int(class_id),
        teacher_name=teacher_name,
        room_name=room_name,
        day_name=day_name,
        period_no=int(period_no),
        academic_year=current_year,
        exclude_id=exclude_id_int
    )
    
    return JsonResponse({"status": "success", "conflicts": conflicts})


@permission_required("timetable.manage")
@transaction.atomic
def api_save_slot(request):
    role = normalized_role(request.user)
    if role == 'Teacher':
        return JsonResponse({"status": "error", "message": "Permission denied"}, status=403)
    """
    POST API to save a dragged slot modification or assignment.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST method required"}, status=400)
        
    timetable_id = request.POST.get('timetable_id')
    class_id = request.POST.get('class_id')
    day_name = request.POST.get('day_name')
    period_no = request.POST.get('period_no')
    subject_id = request.POST.get('subject_id')
    teacher_name = request.POST.get('teacher_name')
    room_name = request.POST.get('room_name')
    
    current_year = active_academic_year()
    p_config = get_object_or_404(TimetablePeriodConfig, period_no=int(period_no))
    
    subject_name = ""
    if subject_id:
        subj = get_object_or_404(Subject, subject_id=int(subject_id))
        subject_name = subj.subject_name
        
    day_orders = {"Monday": 1, "Tuesday": 2, "Wednesday": 3, "Thursday": 4, "Friday": 5}
    day_order = day_orders.get(day_name, 1)

    if timetable_id:
        # Update existing
        entry = get_object_or_404(ClassTimetableEntry, timetable_id=int(timetable_id))
        if entry.is_locked:
            return JsonResponse({"status": "error", "message": "This slot is locked and cannot be moved."}, status=400)
            
        entry.day_name = day_name
        entry.day_order = day_order
        entry.period_no = int(period_no)
        entry.start_time = p_config.start_time
        entry.end_time = p_config.end_time
        if subject_id:
            entry.subject_id = int(subject_id)
            entry.subject_name = subject_name
        if teacher_name is not None:
            entry.teacher_name = teacher_name
        if room_name is not None:
            entry.room_name = room_name
        entry.save()
    else:
        # Create new
        ClassTimetableEntry.objects.create(
            class_id=int(class_id),
            academic_year=current_year,
            day_name=day_name,
            day_order=day_order,
            period_no=int(period_no),
            start_time=p_config.start_time,
            end_time=p_config.end_time,
            subject_id=int(subject_id) if subject_id else None,
            subject_name=subject_name,
            teacher_name=teacher_name,
            room_name=room_name,
            generated_at=timezone.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        
    return JsonResponse({"status": "success"})


@permission_required("timetable.manage")
def api_toggle_lock(request):
    role = normalized_role(request.user)
    if role == 'Teacher':
        return JsonResponse({"status": "error", "message": "Permission denied"}, status=403)
    """
    POST API to lock or unlock a specific timetable slot.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST method required"}, status=400)
        
    timetable_id = request.POST.get('timetable_id')
    is_locked = request.POST.get('is_locked') == 'true'
    
    entry = get_object_or_404(ClassTimetableEntry, timetable_id=int(timetable_id))
    entry.is_locked = is_locked
    entry.save()
    
    return JsonResponse({"status": "success", "is_locked": entry.is_locked})


@permission_required("timetable.view")
def export_pdf(request):
    """
    Exports the current grid view as a landscape PDF.
    """
    view_type = request.GET.get('type', 'class')
    target = request.GET.get('target', '')
    
    role = normalized_role(request.user)
    if role == 'Teacher':
        allowed_class_ids = assigned_classes_for_teacher(request.user)
        if not allowed_class_ids or not target or int(target) not in allowed_class_ids:
            return HttpResponse("Forbidden: You are not allowed to export this timetable.", status=403)
        view_type = 'class'
        
    current_year = active_academic_year()
    s_settings = school_settings()
    
    title = ""
    subtitle = f"Academic Year: {current_year}"
    
    timetable_rows = []
    periods = []
    
    if view_type == 'class' and target:
        class_obj = get_object_or_404(SchoolClass, class_id=target)
        title = f"Weekly Class Timetable - {class_obj.class_name}"
        # Convert DB models to standard dict layout for PDF Generator
        raw_rows, periods = get_grid_data('class', int(target), current_year)
        timetable_rows = raw_rows
    elif view_type == 'teacher' and target:
        title = f"Weekly Teacher Timetable - {target}"
        raw_rows, periods = get_grid_data('teacher', target, current_year)
        timetable_rows = raw_rows
    elif view_type == 'room' and target:
        title = f"Weekly Room Timetable - {target}"
        raw_rows, periods = get_grid_data('room', target, current_year)
        timetable_rows = raw_rows
        
    if not timetable_rows:
        return HttpResponse("No timetable entries found to export.", status=404)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Timetable_{target}.pdf"'
    
    TimetablePDFGenerator.generate_timetable_pdf(
        title=title,
        subtitle=subtitle,
        timetable_rows=timetable_rows,
        periods=periods,
        school_settings=s_settings,
        response_stream=response
    )
    return response


@permission_required("timetable.view")
def export_excel(request):
    """
    Exports the current grid view as an Excel (.xlsx) file.
    """
    view_type = request.GET.get('type', 'class')
    target = request.GET.get('target', '')
    
    role = normalized_role(request.user)
    if role == 'Teacher':
        allowed_class_ids = assigned_classes_for_teacher(request.user)
        if not allowed_class_ids or not target or int(target) not in allowed_class_ids:
            return HttpResponse("Forbidden: You are not allowed to export this timetable.", status=403)
        view_type = 'class'
        
    current_year = active_academic_year()
    
    title = "Weekly Timetable"
    timetable_rows = []
    periods = []
    
    if view_type == 'class' and target:
        class_obj = get_object_or_404(SchoolClass, class_id=target)
        title = f"Timetable_{class_obj.class_name}_{current_year}"
        timetable_rows, periods = get_grid_data('class', int(target), current_year)
    elif view_type == 'teacher' and target:
        title = f"Timetable_Teacher_{target}_{current_year}"
        timetable_rows, periods = get_grid_data('teacher', target, current_year)
    elif view_type == 'room' and target:
        title = f"Timetable_Room_{target}_{current_year}"
        timetable_rows, periods = get_grid_data('room', target, current_year)
        
    if not timetable_rows:
        return HttpResponse("No timetable data found to export.", status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Timetable"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="0F766E")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    day_font = Font(name="Calibri", size=10, bold=True, color="1E293B")
    cell_font = Font(name="Calibri", size=9)
    break_font = Font(name="Calibri", size=10, bold=True, italic=True, color="475569")
    
    teal_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    break_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    slot_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    
    # Write title
    ws.merge_cells('A1:I1')
    ws['A1'] = title.replace("_", " ")
    ws['A1'].font = title_font
    ws['A1'].alignment = align_left
    ws.row_dimensions[1].height = 30
    
    ws.append([]) # Row 2 is blank
    
    # Headers
    headers = ["Day"]
    for p in periods:
        p_label = p.label or (f"Period {p.period_no}" if p.period_type == 'Lesson' else p.period_type)
        headers.append(f"{p.start_time} - {p.end_time}\n({p_label})")
        
    ws.append(headers)
    ws.row_dimensions[3].height = 28
    
    # Style headers
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = teal_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Rows
    for row_idx, row in enumerate(timetable_rows):
        excel_row = [row['day_name']]
        for cell in row['cells']:
            if cell.get('is_break'):
                excel_row.append(cell.get('label', '').upper())
            elif cell.get('is_empty'):
                excel_row.append("-")
            else:
                subj = cell.get('subject_name', '')
                teacher = cell.get('teacher_name', '') or ''
                room = cell.get('room_name', '') or ''
                extra = cell.get('extra_label', '')
                
                parts = [subj]
                if extra:
                    parts.append(extra)
                if teacher and teacher != '-':
                    parts.append(teacher)
                if room and room != '-':
                    parts.append(f"({room})")
                
                excel_row.append("\n".join(parts))
                
        ws.append(excel_row)
        current_row_num = row_idx + 4
        ws.row_dimensions[current_row_num].height = 55
        
        # Style cells
        ws.cell(row=current_row_num, column=1).font = day_font
        ws.cell(row=current_row_num, column=1).alignment = align_center
        ws.cell(row=current_row_num, column=1).border = thin_border
        
        for col_idx, cell_data in enumerate(row['cells']):
            col_num = col_idx + 2
            cell = ws.cell(row=current_row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = align_center
            
            p_config = periods[col_idx]
            if p_config.period_type in ['Break', 'Lunch']:
                cell.font = break_font
                cell.fill = break_fill
            else:
                cell.font = cell_font
                if not cell_data.get('is_empty'):
                    cell.fill = slot_fill
                    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for col_num in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
    wb.save(response)
    return response


# --- Rooms CRUD Views ---
@permission_required("timetable.manage")
def room_list(request):
    rooms = Room.objects.all().order_by('room_name')
    return render(request, "timetable/room_list.html", {"rooms": rooms, "settings": school_settings()})

@permission_required("timetable.manage")
def room_new(request):
    form = RoomForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("timetable:room_list")
    return render(request, "timetable/room_form.html", {"form": form, "title": "New Room", "settings": school_settings()})

@permission_required("timetable.manage")
def room_edit(request, room_id):
    room = get_object_or_404(Room, room_id=room_id)
    form = RoomForm(request.POST or None, instance=room)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("timetable:room_list")
    return render(request, "timetable/room_form.html", {"form": form, "title": "Edit Room", "settings": school_settings()})

@permission_required("timetable.manage")
def room_delete(request, room_id):
    room = get_object_or_404(Room, room_id=room_id)
    if request.method == "POST":
        room.delete()
    return redirect("timetable:room_list")


# --- SubjectAllocations CRUD Views ---
@permission_required("subject_allocations.manage")
def allocation_list(request):
    allocations = SubjectAllocation.objects.all().order_by('school_class__class_name', 'subject__subject_name')
    return render(request, "timetable/allocation_list.html", {"allocations": allocations, "settings": school_settings()})

@permission_required("subject_allocations.manage")
def allocation_new(request):
    form = SubjectAllocationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("timetable:allocation_list")
    return render(request, "timetable/allocation_form.html", {"form": form, "title": "New Subject Allocation", "settings": school_settings()})

@permission_required("subject_allocations.manage")
def allocation_edit(request, allocation_id):
    alloc = get_object_or_404(SubjectAllocation, allocation_id=allocation_id)
    form = SubjectAllocationForm(request.POST or None, instance=alloc)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("timetable:allocation_list")
    return render(request, "timetable/allocation_form.html", {"form": form, "title": "Edit Subject Allocation", "settings": school_settings()})

def sync_subject_allocation_to_registry(allocation):
    import datetime
    from subject_management.models import Subject as RegistrySubject, TeacherSubjectAllocation
    from academic_structure.models import AcademicYear, AcademicTerm, Form, Stream
    
    # 1. Resolve User
    user = allocation.teacher.user
    
    # 2. Resolve Subject
    legacy_code = allocation.subject.subject_code
    reg_sub = RegistrySubject.objects.filter(code=legacy_code).first()
    if not reg_sub:
        reg_sub = RegistrySubject.objects.filter(name__iexact=allocation.subject.subject_name).first()
    
    # 3. Resolve Form and Stream
    class_name = allocation.school_class.class_name
    parts = class_name.strip().split()
    form_val = None
    stream_val = None
    if len(parts) >= 2:
        if parts[0].lower() == "form":
            try:
                form_num = int(parts[1])
                form_val = Form.objects.filter(form_number=form_num).first()
                stream_name = " ".join(parts[2:])
                stream_val = Stream.objects.filter(name__iexact=stream_name).first()
            except ValueError:
                pass
    
    if not form_val:
        form_val = Form.objects.first()
    if not stream_val:
        stream_val = Stream.objects.first()
        
    active_year = AcademicYear.objects.filter(is_active=True).first()
    active_term = AcademicTerm.objects.filter(is_active=True).first()
    
    if not active_year:
        active_year, _ = AcademicYear.objects.get_or_create(year=allocation.school_class.academic_year, is_active=True)
    if not active_term:
        active_term, _ = AcademicTerm.objects.get_or_create(
            academic_year=active_year,
            term_number=2,
            defaults={
                "start_date": datetime.date(2026, 5, 1),
                "end_date": datetime.date(2026, 8, 31),
                "is_active": True
            }
        )
        
    if reg_sub and form_val and stream_val:
        TeacherSubjectAllocation.objects.update_or_create(
            teacher=user,
            subject=reg_sub,
            academic_year=active_year,
            academic_term=active_term,
            form=form_val,
            stream=stream_val
        )

def delete_subject_allocation_from_registry(allocation):
    from subject_management.models import Subject as RegistrySubject, TeacherSubjectAllocation
    from academic_structure.models import Form, Stream
    
    user = allocation.teacher.user
    legacy_code = allocation.subject.subject_code
    reg_sub = RegistrySubject.objects.filter(code=legacy_code).first()
    if not reg_sub:
        reg_sub = RegistrySubject.objects.filter(name__iexact=allocation.subject.subject_name).first()
        
    class_name = allocation.school_class.class_name
    parts = class_name.strip().split()
    form_val = None
    stream_val = None
    if len(parts) >= 2:
        if parts[0].lower() == "form":
            try:
                form_num = int(parts[1])
                form_val = Form.objects.filter(form_number=form_num).first()
                stream_name = " ".join(parts[2:])
                stream_val = Stream.objects.filter(name__iexact=stream_name).first()
            except ValueError:
                pass
                
    if user and reg_sub and form_val and stream_val:
        TeacherSubjectAllocation.objects.filter(
            teacher=user,
            subject=reg_sub,
            form=form_val,
            stream=stream_val
        ).delete()

@permission_required("subject_allocations.manage")
def allocation_delete(request, allocation_id):
    alloc = get_object_or_404(SubjectAllocation, allocation_id=allocation_id)
    if request.method == "POST":
        try:
            delete_subject_allocation_from_registry(alloc)
        except Exception as e:
            print(f"Error deleting subject allocation from registry: {e}")
        alloc.delete()
    return redirect("timetable:allocation_list")


# --- TeacherAvailability CRUD Views ---
@permission_required("timetable.manage")
def availability_list(request):
    availabilities = TeacherAvailability.objects.all().order_by('teacher__full_name')
    return render(request, "timetable/availability_list.html", {"availabilities": availabilities, "settings": school_settings()})

@permission_required("timetable.manage")
def availability_new(request):
    form = TeacherAvailabilityForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("timetable:availability_list")
    return render(request, "timetable/availability_form.html", {"form": form, "title": "New Teacher Availability", "settings": school_settings()})

@permission_required("timetable.manage")
def availability_edit(request, availability_id):
    avail = get_object_or_404(TeacherAvailability, availability_id=availability_id)
    form = TeacherAvailabilityForm(request.POST or None, instance=avail)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("timetable:availability_list")
    return render(request, "timetable/availability_form.html", {"form": form, "title": "Edit Teacher Availability", "settings": school_settings()})

@permission_required("timetable.manage")
def availability_delete(request, availability_id):
    avail = get_object_or_404(TeacherAvailability, availability_id=availability_id)
    if request.method == "POST":
        avail.delete()
    return redirect("timetable:availability_list")


# --- TimetablePeriodConfig CRUD Views ---
@permission_required("timetable.manage")
def period_config_list(request):
    configs = TimetablePeriodConfig.objects.all().order_by('period_no')
    return render(request, "timetable/period_config_list.html", {"configs": configs, "settings": school_settings()})

@permission_required("timetable.manage")
def period_config_new(request):
    form = TimetablePeriodConfigForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("timetable:period_config_list")
    return render(request, "timetable/period_config_form.html", {"form": form, "title": "New Period Configuration", "settings": school_settings()})

@permission_required("timetable.manage")
def period_config_edit(request, config_id):
    config = get_object_or_404(TimetablePeriodConfig, config_id=config_id)
    form = TimetablePeriodConfigForm(request.POST or None, instance=config)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("timetable:period_config_list")
    return render(request, "timetable/period_config_form.html", {"form": form, "title": "Edit Period Configuration", "settings": school_settings()})

@permission_required("timetable.manage")
def period_config_delete(request, config_id):
    config = get_object_or_404(TimetablePeriodConfig, config_id=config_id)
    if request.method == "POST":
        config.delete()
    return redirect("timetable:period_config_list")
